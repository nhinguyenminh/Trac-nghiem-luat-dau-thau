import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
INPUT_FILES = [
    ROOT / "data" / "390-cau-hoi-luat-dau-thau.xlsx",
    ROOT / "data" / "50-cau-bo-sung.xlsx",
]
OUTPUT_JSON = ROOT / "public" / "questions.json"

ANSWER_MAP = {"A": 0, "B": 1, "C": 2, "D": 3}

RULES: list[tuple[str, list[str]]] = [
    (
        "Đấu thầu theo FTA",
        [
            r"\bcptpp\b",
            r"\bevfta\b",
            r"\bukvfta\b",
            r"hiệp định",
            r"mua sắm chính phủ",
            r"cam kết",
            r"nghị định\s+số\s+95/2020",
        ],
    ),
    (
        "Đấu thầu qua mạng",
        [
            r"hệ thống mạng đấu thầu",
            r"đấu thầu qua mạng",
            r"e-hsmt",
            r"e-hsdt",
            r"chào giá trực tuyến",
            r"mua sắm trực tuyến",
            r"tài khoản nghiệp vụ",
            r"văn bản điện tử",
            r"đăng tải",
        ],
    ),
    (
        "Đấu thầu quốc tế",
        [
            r"đấu thầu quốc tế",
            r"nhà thầu nước ngoài",
            r"ngôn ngữ sử dụng",
            r"hồ sơ mời thầu.*quốc tế",
            r"quốc tế",
        ],
    ),
    (
        "Giải quyết kiến nghị",
        [
            r"kiến nghị",
            r"khiếu nại",
            r"giải quyết kiến nghị",
            r"kết quả lựa chọn nhà thầu",
        ],
    ),
    (
        "Thanh tra, kiểm tra & xử lý vi phạm",
        [
            r"thanh tra",
            r"kiểm tra hoạt động đấu thầu",
            r"xử lý vi phạm",
            r"cấm tham gia hoạt động đấu thầu",
            r"thời hiệu",
        ],
    ),
    (
        "Mua sắm tập trung",
        [
            r"mua sắm tập trung",
            r"đơn vị mua sắm tập trung",
            r"thỏa thuận khung",
            r"danh mục.*mua sắm tập trung",
        ],
    ),
    (
        "Hợp đồng",
        [
            r"hợp đồng",
            r"thanh toán hợp đồng",
            r"bảo đảm thực hiện hợp đồng",
            r"hoàn thiện.*ký kết hợp đồng",
            r"điều chỉnh hợp đồng",
        ],
    ),
    (
        "Kế hoạch lựa chọn nhà thầu",
        [
            r"kế hoạch lựa chọn nhà thầu",
            r"\bkhlcnt\b",
            r"giá gói thầu",
            r"phê duyệt kế hoạch",
        ],
    ),
    (
        "Hình thức lựa chọn nhà thầu",
        [
            r"hình thức lựa chọn nhà thầu",
            r"phương thức lựa chọn nhà thầu",
            r"đấu thầu rộng rãi",
            r"đấu thầu hạn chế",
            r"chỉ định thầu",
            r"chào hàng cạnh tranh",
            r"mua sắm trực tiếp",
            r"sơ tuyển",
            r"đặt hàng",
        ],
    ),
    (
        "HSMT & tiêu chuẩn đánh giá",
        [
            r"hồ sơ mời thầu",
            r"hồ sơ yêu cầu",
            r"tiêu chuẩn đánh giá",
            r"hồ sơ mời quan tâm",
            r"hồ sơ mời sơ tuyển",
            r"chương\s*v",
        ],
    ),
    (
        "Đánh giá HSDT & xét thầu",
        [
            r"đánh giá hồ sơ dự thầu",
            r"hồ sơ dự thầu",
            r"xếp hạng",
            r"xét thầu",
            r"sửa lỗi",
            r"hiệu chỉnh sai lệch",
            r"giá dự thầu",
            r"năng lực kinh nghiệm",
        ],
    ),
    (
        "Tư cách & cạnh tranh",
        [
            r"bảo đảm cạnh tranh",
            r"tư cách",
            r"độc lập với",
            r"năng lực",
            r"nghĩa vụ nộp thuế",
            r"bảo đảm dự thầu",
        ],
    ),
    (
        "Chủ thể tham gia đấu thầu",
        [
            r"chủ đầu tư",
            r"người có thẩm quyền",
            r"tổ chuyên gia",
            r"bên mời thầu",
            r"trách nhiệm",
        ],
    ),
    (
        "Mua sắm hàng hóa",
        [
            r"mua sắm hàng hóa",
            r"hàng hóa",
            r"thuốc",
            r"hóa chất",
            r"vật tư xét nghiệm",
            r"thiết bị y tế",
        ],
    ),
    (
        "Tổng quan & phạm vi áp dụng",
        [
            r"đấu thầu là gì",
            r"phạm vi áp dụng",
            r"đối tượng áp dụng",
            r"phạm vi điều chỉnh",
            r"quy định pháp luật về đấu thầu",
        ],
    ),
]


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", text.lower()).strip()


def split_options(options_text: str) -> list[str]:
    options: list[str] = []
    for line in str(options_text or "").splitlines():
        line = line.strip()
        if not line:
            continue
        if re.match(r"^[A-Da-d][\).:\-\s]", line):
            line = re.sub(r"^[A-Da-d][\).:\-\s]+", "", line).strip()
        options.append(line)
    return options


def classify(question: str, options_text: str, answer: str, fallback_category: str) -> str:
    options = split_options(options_text)
    answer_index = ANSWER_MAP.get(str(answer).strip().upper(), -1)
    answer_text = options[answer_index] if 0 <= answer_index < len(options) else ""
    corpus = normalize(f"{question} {' '.join(options)} {answer_text}")

    best_category = fallback_category.strip() if fallback_category else "Tổng quan & phạm vi áp dụng"
    best_score = 0

    for category, patterns in RULES:
        score = 0
        for pattern in patterns:
            if re.search(pattern, corpus):
                score += 1
        if score > best_score:
            best_score = score
            best_category = category

    if best_score == 0:
        # Fallback heuristic for unresolved rows.
        if "trường hợp" in corpus or "xử lý" in corpus:
            return "Tình huống thực tiễn & nâng cao"
        return best_category

    # Situational overlay: keep domain category unless explicitly a handling scenario.
    if (
        "trường hợp" in corpus
        and ("xử lý" in corpus or "như thế nào" in corpus)
        and best_category
        not in {"Giải quyết kiến nghị", "Thanh tra, kiểm tra & xử lý vi phạm", "Hợp đồng"}
    ):
        return "Tình huống thực tiễn & nâng cao"

    return best_category


def update_workbook(path: Path) -> tuple[int, int]:
    wb = load_workbook(path)
    ws = wb.active

    updated = 0
    total = 0

    for row in ws.iter_rows(min_row=2):
        stt = row[0].value
        if stt is None:
            continue

        total += 1
        current_category = str(row[1].value or "").strip()
        question = str(row[2].value or "").strip()
        options_text = str(row[3].value or "")
        answer = str(row[4].value or "")

        next_category = classify(question, options_text, answer, current_category)
        if next_category != current_category:
            row[1].value = next_category
            updated += 1

    wb.save(path)
    return updated, total


def regenerate_json() -> None:
    questions = []
    next_id = 1

    for input_path in INPUT_FILES:
        wb = load_workbook(input_path, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            stt, category, question, options_text, answer = row[0], row[1], row[2], row[3], row[4]
            if stt is None:
                continue

            options = split_options(str(options_text or ""))
            if not options:
                continue

            answer_index = ANSWER_MAP.get(str(answer).strip().upper(), 0)

            questions.append(
                {
                    "id": next_id,
                    "category": str(category or "").strip(),
                    "question": str(question or "").strip(),
                    "options": options,
                    "answer": answer_index,
                }
            )
            next_id += 1

    OUTPUT_JSON.write_text(json.dumps(questions, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    report = []
    for path in INPUT_FILES:
        updated, total = update_workbook(path)
        report.append((path.name, updated, total))

    regenerate_json()

    for name, updated, total in report:
        print(f"{name}: updated {updated}/{total} rows")
    print(f"Regenerated {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
