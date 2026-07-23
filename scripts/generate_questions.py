import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
INPUT_FILES = [
    ROOT / "data" / "390-cau-hoi-luat-dau-thau.xlsx",
]
OUTPUT_PATH = ROOT / "public" / "questions.json"

answer_map = {"A": 0, "B": 1, "C": 2, "D": 3}

questions = []
next_id = 1

for input_path in INPUT_FILES:
    if not input_path.exists():
        print(f"Skipping missing file: {input_path}")
        continue

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        _, category, question, options_text, answer = row[0], row[1], row[2], row[3], row[4]
        if row[0] is None:
            continue

        options = []
        if isinstance(options_text, str):
            for line in options_text.splitlines():
                line = line.strip()
                if not line:
                    continue
                if line and line[0].upper() in answer_map:
                    line = line[2:].strip() if len(line) > 2 else ""
                options.append(line)

        if not options:
            continue

        answer_index = answer_map.get(str(answer).strip().upper())
        if answer_index is None:
            answer_index = 0

        questions.append({
            "id": next_id,
            "category": str(category).strip() if category is not None else "",
            "question": str(question).strip(),
            "options": options,
            "answer": answer_index,
        })
        next_id += 1

OUTPUT_PATH.write_text(json.dumps(questions, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {len(questions)} questions to {OUTPUT_PATH}")
